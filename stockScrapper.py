import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import math as m
import os
import csv
import time
from decimal import Decimal as dc
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font
##################################ADDITIONAL FUNCTIONS###########################
def rgb_to_hex(rgb):
    return "{:02x}{:02x}{:02x}".format(*rgb)
##################################STOCK FUNCTIONS################################
#?????????????????????????????
def get_perc_change(n_days, history_cls):
    temp_array = history_cls[:n_days]
    first_item = temp_array[0]
    last_item = temp_array[-1]
    perc_change = (first_item - last_item)/last_item * 100
    return perc_change
#???????????????????????????
def get_volas(n_days, history_cls):
    temp_price = history_cls[:n_days]
    avg_price = sum(temp_price) / n_days
    diff_cls = []
    for i in range(n_days):
        diff = temp_price[i] - avg_price
        diff_cls.append(diff)
    squared_diff = []
    for i in range(n_days):
        diff_cls[i] = diff_cls[i]**2
        
    variance = sum(diff_cls) / n_days
    std_dev = m.sqrt(variance)
    return std_dev

#????????????????????????????
def get_avg_vol(n_days,history_vol):
    temp_sum = 0
    for i in range(n_days):
        temp_sum += history_vol[i]
    average_volume = temp_sum / n_days
    return average_volume
#????????????????????????????
def get_ma(n_days,history_cls):
    temp_sum = 0
    for i in range(n_days):
        temp_sum += history_cls[i]
    average_volume = temp_sum / n_days
    return average_volume
#????????????????????????????
def calculate_display_ma(ma5, ma10, ma20, ma50, ma100, ma200,stock):
    symbols = []
    output_const = 0
    # Compare MA10 with MA5
    if stock > ma5:
        symbols.append(">")
        output_const += 1
    elif stock == ma5:
        symbols.append("=")
        output_const += 0
    else:
        symbols.append("<")

    # Compare MA20 with MA10
    if stock > ma10:
        symbols.append(">")
        output_const += 2
    elif stock == ma10:
        symbols.append("=")
        output_const += 1
    else:
        symbols.append("<")

    # Compare MA50 with MA20
    if stock > ma20:
        symbols.append(">")
        output_const += 5
    elif stock == ma20:
        symbols.append("=")
        output_const += 4
    else:
        symbols.append("<")

    # Compare MA100 with MA50
    if stock > ma50:
        symbols.append(">")
        output_const += 4
    elif stock == ma50:
        symbols.append("=")
        output_const += 3
    else:
        symbols.append("<")

    # Compare MA200 with MA100
    if stock > ma100:
        symbols.append(">")
        output_const += 3
    elif stock == ma100:
        symbols.append("=")
        output_const += 2
    else:
        symbols.append("<")
        
    if stock > ma200:
        symbols.append(">")
        output_const += 6
    elif stock == ma200:
        symbols.append("=")
        output_const += 5
    else:
        symbols.append("<")

    return symbols, output_const
#????????????????????????????
def analyze_bound_stock(n_days,history_cls,history_vol):
    # print("closing",closing_prices)
    output_const = 0.0
    i = 0
    is_avg_check = False
    current_array = []
    multiplier = n_days/2
    temp_price = history_cls[:n_days]
    temp_vol = history_vol[:n_days]
    avg_price = sum(temp_price) / n_days
    avg_vol = sum(temp_vol) / n_days
    upper_bound_avg_vol = avg_vol + avg_vol/multiplier
    lower_bound_avg_vol = avg_vol - avg_vol/multiplier
    # upper_bound = avg_price + (avg_price / multiplier)
    lower_bound = avg_price - (avg_price / multiplier)
    while not is_avg_check and i< n_days:
        current_array.append(history_cls[i])
        is_low_vol = True if lower_bound_avg_vol < history_vol[i] < upper_bound_avg_vol else False
        # print("lower vs next vs upper:",lower_bound, "<", history_cls[i+1], "<", upper_bound)
        # print("history vs avg:",history_vol[i], "<", avg_vol)
        if (lower_bound < history_cls[i+1]) and is_low_vol:
            i += 1
            output_const += 1
            # print("next iteration",n_days)
        else:
            print("done iter",n_days)
            is_avg_check = True
    # output_const = output_const * (market_cap/1000000000)
    return output_const
    
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
##################################MAIN################################
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
#000000#000000#000000#000000#000000#000000#000000#000000#000000#000000
# Step 1: Create empty Excel file and debug text file
current_date = datetime.now().strftime("%d-%m-%Y")
excel_filename = f"stock_data_{current_date}.xlsx"
debug_filename = f"debug_stock_scrapper_{current_date}.txt"
temp_color = (0,0,0)
# Create empty Excel file
df_empty = pd.DataFrame()
df_empty.to_excel(excel_filename, index=False)

# Create empty debug text file
with open(debug_filename, 'w') as debug_file:
    debug_file.write("Debug Log - Stock Scrapper\n\n")

# Print header
header = [
    'Stock',
    'Buy Analysis Result',
    'Buy Analysis 5 Days',
    'Buy Analysis 20 Days',
    'Buy Analysis 50 Days',
    'MA Analysis',
    '-',
    '=',
    'MA5',
    'MA5 Symbol',
    'MA10',
    'MA10 Symbol',
    'MA20',
    'MA20 Symbol',
    'MA50',
    'MA50 Symbol',
    'MA100',
    'MA100 Symbol',
    'MA200',
    'MA200 Symbol',
    '3D Change%',
    '5D Change%',
    '20D Change%',
    'Yesterday Closing Price',
    'Current Price',
    'Volatility in 3 Day',
    'Volatility in 5 Days',
    'Volatility in 20 Days',
    'Market Cap Value',
    'Volume Average in 5 Days',
    'Volume Average in 20 Days',
    'Volume Average in 50 Days',
    'Volume Average in 100 Days',
]


# Print header to debug file
with open(debug_filename, 'a') as debug_file:
    debug_file.write("Header in Excel File:\n")
    debug_file.write("\t".join(header) + "\n")

# Append header to Excel file
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    pd.DataFrame(columns=header).to_excel(writer, index=False, sheet_name='Sheet1')

# Step 2: Read stock list from CSV and add ".JK" suffix
stocks = []
with open('stocklist.csv', 'r') as csvfile:
    reader = csv.reader(csvfile)
    for row in reader:
        stock = row[0] + ".JK"
        stocks.append(stock)

# #000000
# stocks = stocks[:30]        
        
#####################################THE THINGY#########################################
# Step 3-6: Fetch and append data for each stock
for stock in stocks:
    try:
        stock_data = yf.Ticker(stock)

        # Fetch required stock information
        info = stock_data.info
        history = stock_data.history(period='3y', interval="1d")
        # print("stock history\n",history)
        history_cls = history['Close'].tolist()
        # current_cls = history_cls.pop() #!!!! bila butuh
        history_cls.reverse()
        # print("cls history\n",history_cls)
        history_vol = history['Volume'].tolist()
        # current_vol = history_vol.pop() #!!!! bila butuh
        history_vol.reverse()
        # print("vol history\n",history_vol)
        # history = history.pop()
        # history.reverse()
        # print("stock history",history)
        # print(type(last_n_closing_prices))
        perc3 = get_perc_change(3,history_cls)#
        perc5 = get_perc_change(5,history_cls)#
        perc20 = get_perc_change(20,history_cls)#
        #volatility
        vola3 = get_volas(3,history_cls)#
        vola5 = get_volas(5,history_cls)#
        vola20 = get_volas(20,history_cls)#
        #avg volume
        vol5= get_avg_vol(5,history_vol)#
        vol20 = get_avg_vol(20,history_vol)#
        vol50 = get_avg_vol(50,history_vol)#
        vol100 = get_avg_vol(100,history_vol)#
        #ma
        ma5 = get_ma(5, history_cls)
        ma10 = get_ma(10, history_cls)
        ma20 = get_ma(20, history_cls)
        ma50 = get_ma(50, history_cls)
        ma100 = get_ma(100, history_cls)
        ma200 = get_ma(200, history_cls)
        display_ma,ma_const = calculate_display_ma(ma5, ma10, ma20, ma50, ma100, ma200,history_cls[0])
        buy_const5= analyze_bound_stock(5,history_cls,history_vol)
        buy_const20= analyze_bound_stock(20,history_cls,history_vol)
        buy_const50= analyze_bound_stock(50,history_cls,history_vol)
        # Create a DataFrame with the required information
        stock_info = pd.DataFrame({
            'Stock': [stock],
            'Buy Analysis Result':[buy_const5 + buy_const20 + buy_const50],
            'Buy Analysis 5 Days': [buy_const5],
            'Buy Analysis 20 Days': [buy_const20],
            'Buy Analysis 50 Days': [buy_const50],
            'MA Analysis': [ma_const],
            '-': 0,
            '=': 0,
            'MA5': [ma5],
            'MA5 Symbol': [display_ma[0]],
            'MA10': [ma10],
            'MA10 Symbol': [display_ma[1]],
            'MA20': [ma20],
            'MA20 Symbol': [display_ma[2]],
            'MA50': [ma50],
            'MA50 Symbol': [display_ma[3]],
            'MA100': [ma100],
            'MA100 Symbol': [display_ma[4]],
            'MA200': [ma200],
            'MA200 Symbol': [display_ma[5]],
            '3D Change%': [perc3],
            '5D Change%': [perc5],
            '20D Change%': [perc20],
            'Yesterday Closing Price': [history_cls[1]],
            'Current Price': [history_cls[0]],
            'Volatility in 3 Day': [vola3],
            'Volatility in 5 Days': [vola5],
            'Volatility in 20 Days': [vola20],
            'Market Cap Value': [info['marketCap']],
            'Volume Average in 5 Days': [vol5],
            'Volume Average in 20 Days': [vol20],
            'Volume Average in 50 Days': [vol50],
            'Volume Average in 100 Days': [vol100],
            })
        print(stock_info)
        
        
        # Append to the Excel file
        wb = load_workbook(excel_filename)
        sheet = wb['Sheet1']
        for index, row in stock_info.iterrows():
            sheet.append(row.tolist())
        wb.save(excel_filename)
        # with pd.ExcelWriter(excel_filename, mode='a', engine='openpyxl') as writer:
        #     stock_info.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
            
        # Log success in debug file
        with open(debug_filename, 'a') as debug_file:
            debug_file.write(f"Successfully processed {stock}.\n")

    except Exception as e:
        # Log errors in debug file
        with open(debug_filename, 'a') as debug_file:
            debug_file.write(f"Error processing {stock}: {str(e)}\n")

    # Add a delay of 3 seconds between requests
    print("done: ", stock)
    time.sleep(3)

#recolor result
wb = load_workbook(excel_filename)
ws = wb.active
col_index = 0
for cell in ws[1]:
    if cell.internal_value == 'Buy Analysis 5 Days':
        col_index = cell.col_idx
        break
for cells_in_row in ws.iter_rows(min_row=2):
    cell_value = cells_in_row[col_index-1].internal_value
    mod_value = max(min(cell_value, 255), 0)
    rgb_color = (0,int(mod_value),0)
    hex_color = rgb_to_hex(rgb_color)
    cells_in_row[col_index-1].font = Font(hex_color)
wb.save(excel_filename)