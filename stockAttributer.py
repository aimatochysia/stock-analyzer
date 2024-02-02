import openpyxl
from openpyxl.styles import PatternFill

def color_cells(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Select the active sheet
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        # Access the cell in the sixth column
        cell = row[0]
        
        # Get the value of the cell
        value = cell.value
        
        # Define the fill color based on the logic
        if value > 1:
            fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
        elif 1 <= value <= 1:
            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
        else:
            fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        
        # Apply the fill color to the cell
        cell.fill = fill
    wb.save(file_path)
    print("done 1")
    
    for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        # Access the cell in the sixth column
        cell = row[0]
        
        # Get the value of the cell
        value = cell.value
        
        # Define the fill color based on the logic
        if value > 1:
            fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
        elif 1 <= value <= 1:
            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
        else:
            fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        
        # Apply the fill color to the cell
        cell.fill = fill
    wb.save(file_path)
    print("done 2")
    
    for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        # Access the cell in the sixth column
        cell = row[0]
        
        # Get the value of the cell
        value = cell.value
        
        # Define the fill color based on the logic
        if value > 1:
            fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
        elif 1 <= value <= 1:
            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
        else:
            fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        
        # Apply the fill color to the cell
        cell.fill = fill
    # Save the modified workbook
    wb.save(file_path)
    print("done 3")
    
    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=5):
        # Access the cell in the sixth column
        cell = row[0]
        
        # Get the value of the cell
        value = cell.value
        
        # Define the fill color based on the logic
        if value > 1:
            fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
        elif 1 <= value <= 1:
            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
        else:
            fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        
        # Apply the fill color to the cell
        cell.fill = fill
        
    # Save the modified workbook
    wb.save(file_path)
    print("done 4")
    
    # Iterate through each row starting from the second row (excluding the header)
    for row in sheet.iter_rows(min_row=2, min_col=6, max_col=6):
        # Access the cell in the sixth column
        cell = row[0]
        
        # Get the value of the cell
        value = cell.value
        
        # Define the fill color based on the logic
        if value > 15:
            fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
        elif 10 <= value <= 15:
            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
        else:
            fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
        
        # Apply the fill color to the cell
        cell.fill = fill
    
    # Save the modified workbook
    wb.save(file_path)
    print("done 5")
    
    
# Specify the path to your Excel file
excel_file_path = 'stock_data_31-01-2024.xlsx'

# Call the function to recolor the cells
color_cells(excel_file_path)
print("done")
